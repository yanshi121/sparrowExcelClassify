import os.path
import statistics
import xlsxwriter as xw
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side


class DealExcel(object):
    def __init__(self, data, choose_data, choose_sheet_data):
        self.excel_data = data
        self.excel_path = r"./static/uploads/split_data"
        self.choose_data = choose_data
        self.choose_sheet_data = choose_sheet_data

    def to_excel(self):
        for name, name_value in self.excel_data.items():
            if name != "":
                data = {}
                workbook = xw.Workbook(os.path.join(self.excel_path, f"{name}.xlsx"))
                for sheet_name, sheet_value in name_value.items():
                    worksheet = workbook.add_worksheet(sheet_name)
                    worksheet.activate()
                    i = int(self.choose_data[sheet_name]['start'])
                    data[sheet_name] = {}
                    data['num'] = len(sheet_value)
                    for x in range(len(sheet_value[0])):
                        data[sheet_name][str(x)] = []
                    for j in range(len(sheet_value)):
                        for x1 in range(len(sheet_value[j])):
                            data[sheet_name][str(x1)].append(sheet_value[j][x1])
                        row = "A" + str(i)
                        worksheet.write_row(row, sheet_value[j])
                        i += 1
                workbook.close()
                workbook = load_workbook(os.path.join(self.excel_path, f"{name}.xlsx"))
                for sheet_name, sheet_value in self.choose_sheet_data.items():
                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for index, value in sheet_value.items():
                            sheet.merge_cells(f"{value['startMerge']}:{value['endMerge']}")
                            if value['selectValue'] == 'sign':
                                sheet[value['startMerge']] = name
                            elif value['selectValue'] == 'num':
                                sheet[value['startMerge']] = data["num"]
                            elif value['selectValue'] == 'sum':
                                sheet[value['startMerge']] = sum(data[sheet_name][(value['sum'])])
                            elif value['selectValue'] == 'mean':
                                sheet[value['startMerge']] = sum(data[sheet_name][value['mean']]) / data["num"]
                            elif value['selectValue'] == 'median':
                                sheet[value['startMerge']] = statistics.median(sorted(data[sheet_name][value['median']]))
                            elif value['selectValue'] == 'min':
                                sheet[value['startMerge']] = min(data[sheet_name][value['min']])
                            elif value['selectValue'] == 'max':
                                sheet[value['startMerge']] = max(data[sheet_name][value['max']])
                            elif value['selectValue'] == 'userDefined':
                                sheet[value['startMerge']] = value['userDefined']
                            cell = sheet[value['startMerge']]
                            cell.font = Font(name=value['fontType'], bold=True, size=int(value['fontSize']))
                            if value['alignment'] == "median":
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            elif value['alignment'] == "left":
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            elif value['alignment'] == "right":
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                workbook.save(os.path.join(self.excel_path, f"{name}.xlsx"))
                workbook.close()
