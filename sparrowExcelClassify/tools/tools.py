import os
import shutil
import zipfile
from openpyxl import load_workbook


def create():
    os.makedirs("./static/uploads/split_data", exist_ok=True)


def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def open_xlsx(path=None):
    if path is None:
        sheet = load_workbook(r'./static/uploads/upload.xlsx')
    else:
        sheet = load_workbook(path)
    return sheet


def return_sheet_name(sheet):
    sheet_name = sheet.sheetnames
    return sheet_name


def get_split_information(data):
    split_information = {}
    sheets = data.get("sheets").split(",")
    for sheet in sheets:
        split_information[sheet] = {}
        split_information[sheet]["start"] = data.get(sheet + "_start")
        split_information[sheet]["classify"] = data.get(sheet + "_classify")
        split_information[sheet]["choose"] = data.get(sheet + "_choose")
    return split_information, sheets


def get_split_choose_information(data):
    split_choose_information = {}
    for k, v in data.items():
        data_list = k.split("_")
        name = data_list[0]
        project = data_list[1]
        value = data_list[2]
        if name in split_choose_information.keys():
            if value in split_choose_information[name].keys():
                split_choose_information[name][value][project] = v
            else:
                split_choose_information[name][value] = {project: v}
        else:
            split_choose_information[name] = {value: {project: v}}
    return split_choose_information


def classify_worksheet(split_information, workbook):
    sheets = workbook.sheetnames
    all_row_data = {}
    for sheet in sheets:
        if split_information.get(sheet).get("choose") == "yes":
            classify = int(split_information[sheet].get("classify")) - 1
            start = int(split_information[sheet].get("start"))
            worksheet = workbook[sheet]
            name_ = ""
            for row in worksheet.iter_rows(min_row=start, values_only=True):
                row_data = list(row)
                name = row_data[classify]
                if name == "" or name is None:
                    name = name_
                elif name[0] == " ":
                    name = name_
                else:
                    name_ = row_data[classify]
                if name in all_row_data.keys():
                    if sheet in all_row_data[name].keys():
                        all_row_data[name][sheet].append(row_data)
                    else:
                        all_row_data[name][sheet] = [row_data]
                else:
                    all_row_data[name] = {sheet: [row_data]}
    return all_row_data


def compress_folder(folder_path=r'./static/uploads/split_data', zip_file_path=r'./static/uploads/output.zip'):
    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                file_in_zip_path = os.path.relpath(file_path, os.path.dirname(folder_path))
                zipf.write(file_path, file_in_zip_path)


